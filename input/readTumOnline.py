#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import division

import sys
import os
PATHS = os.getcwd().split('/')
PROJECT_PATH = ''
for p in PATHS:
    PROJECT_PATH += '%s' % p
    if p == 'examination-scheduling':
        break
sys.path.append(PROJECT_PATH)

from time import time
import random 
import openpyxl

def read_real_data():
    wb = openpyxl.load_workbook('%s\input\Data\TumOnline\Read15S.xlsx' % (PROJECT_PATH))
    ws = wb.get_sheet_by_name('Conflicts')

    maxRow = ws.max_row

    # create number of exams
    n = maxRow-1

    # create Conflict Matrix
    Q = [[int(cell.value) if not cell.value is None else 0 for cell in row]   for row in ws.iter_rows("B2:%s%s" % (openpyxl.utils.get_column_letter(maxRow),maxRow))]



	
    # np.random.seed(kwards.get('tseed', 1))
    # rd.seed(kwards.get('tseed', 1))
    # n, r, p, w = kwards.get('n', 0), kwards.get('r', 0), kwards.get('p', 0), kwards.get('w', ["1", "2", "3", "4", "5", "6", "7"])
    # data = {'n': n, 'r': r, 'p': p}

    # #create possible number of participants, increase probability that number of participants is between 150 and 300
    # num = [i for i in range(10,901)]
    # for times in range(1500):
    #     num.extend([int(i) for i in range(10,150)])
    # for times in range(500):
    #     num.extend([int(i) for i in range(150,301)])

    # # get number of students participating
    # data['s'] = np.random.choice(num, n)

    # # get room capacity from real data
    # data['c'] = np.random.choice(num, r)
    # data['c'] = sorted(data['c'], reverse=True)

    # if kwards.get('locations') == True:
    # 	data['w'] = np.random.choice([["1"], ["2"], ["3"], ["2","3"], ["1","2"], ["1","3"], ["1","2","3"]], n , p=[0.2, 0.1, 0.05, 0.05, 0, 0, 0.6])
    # 	data['location'] = np.random.choice(["1", "2", "3"], r , p=[0.6, 0.35, 0.05])
    
    # # hours between starting day and starting periods are fixed equal to 2
    # data['h'] = [ 2*l for l in range(p)]
    
    # # create a conflict by probybility 1/5
    # data['conflicts'] = defaultdict(list)
    # for i in range(n):
    #     data['conflicts'][i] = [ j for j in range(i+1,n) if rd.random() <= 0.1 ]
    
    # #close some rooms by probability 1/10
    # data['locking_times'] = defaultdict(list)
    # for k in range(r):
    #     data['locking_times'][k] = [ l for l in range(p) if rd.random() <= 0.1 ]
    
    return 


if __name__ == '__main__':
    read_real_data()