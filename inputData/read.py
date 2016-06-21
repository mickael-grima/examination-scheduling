import sys
import os
PATHS = os.getcwd().split('/')
PROJECT_PATH = ''
for p in PATHS:
    PROJECT_PATH += '%s/' % p
    if p == 'examination-scheduling':
        break
sys.path.append(PROJECT_PATH)

from time import time
import random 
import openpyxl


def read_real_data():
    try:
        wb = openpyxl.load_workbook('%s\inputData\Data\TumOnline\Read15S.xlsx' % (PROJECT_PATH))
    except:
        wb = openpyxl.load_workbook('%s/inputData/Data/TumOnline/Read15S.xlsx' % (PROJECT_PATH))
    
    ws = wb.get_sheet_by_name('Conflicts')
    ws2 = wb.get_sheet_by_name('Students')
    ws3 = wb.get_sheet_by_name('Raum')
    
    maxRow = ws.max_row
    maxRow3 = ws3.max_row

    # create number of exams and room
    data = {'n': maxRow-1, 'r': maxRow3-1}

    # create Conflict Matrix
    data['Q'] = [[int(cell.value) if not cell.value is None else 0 for cell in row]   for row in ws.iter_rows("B2:%s%s" % (openpyxl.utils.get_column_letter(maxRow),maxRow))]

    # create number of students participating
    data['s'] = [int(ws2["A%s" % (i)].value) for i in range(1,maxRow)]

    # create room capacities
    data['c'] = sorted([ int(ws3["D%s" % (k)].value) for k in range(2,maxRow3+1)], reverse=True)

    return data


if __name__ == '__main__':
    read_real_data()